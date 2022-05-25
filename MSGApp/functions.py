

def handle_uploaded_file(f):  
    
    #import openpyxl library to handle the excel file
    from openpyxl import load_workbook


    #connect to the database
    #database = MySQLdb.connect (host="localhost" , user="MSG" , passwd="MSG19900" ,db="hwcoredata")
    #cursor = database.cursor()
    from django.db import connection
    cursor = connection.cursor()


    #create the table if not exists
    hubenergy_table = ("CREATE TABLE IF NOT EXISTS hubenergy(id int,date_month date NOT NULL, hub varchar(50) ,hub_id varchar(10),Production decimal(5,2) NOT NULL,Consumption decimal(5,2) NOT NULL, PRIMARY KEY(id) )")
    cursor.execute(hubenergy_table)

    #read excel_sheet
    path='media/uploads/'+ f.name
    wb = load_workbook(path)
    #ws = wb[wb.sheetnames[0]] # Select the first sheet
    ws = wb['Sheet1']# or use the sheetname directly

    rows = ws.max_row # Get the maximum number of rows in the table
    columns = ws.max_column   # Get the maximum number of columns in the table

   
    #insert into the database table
    data_start_line = 2 #ignore headings row and start on row 2
    for row in range(data_start_line, rows + 1):
        print ('###############')
        print ('row '+ str(row) + ' inserted')
        print ('###############')
        
        id = ws.cell(row=row, column=1).value    #month_date
        date_month= ws.cell(row=row, column=2).value    #month_date
        hub = ws.cell(row=row, column=3).value    #hub
        hub_id = ws.cell(row=row, column=4).value    #hub_id
        production = ws.cell(row=row, column=5).value    #production
        consumption = ws.cell(row=row, column=6).value    #consumption
        
        query = """insert into hubenergy (id,date_month, hub,hub_id ,Production , Consumption) values ("%s","%s","%s","%s","%s","%s")"""%( id,date_month, hub,hub_id ,production , consumption)

        #print query
        
        cursor.execute(query)
        connection.commit()
    
    
    

